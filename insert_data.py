import sqlite3
import openpyxl
import sys

def insert_data_from_excel(file_path, table_name):
    # Conexión a la base de datos SQLite
    conn = sqlite3.connect('your_database')
    cursor = conn.cursor()

    # Leer datos desde el archivo Excel
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active

    # Obtener los nombres de las columnas de la primera fila
    column_names = [cell.value for cell in next(sheet.iter_rows())]

    # Iterar sobre las filas del archivo Excel, omitiendo la primera fila
    for row in sheet.iter_rows(min_row=2, values_only=True):
        # Crear una cadena de marcadores de posición para los valores de la fila
        placeholders = ', '.join(['?' for _ in range(len(row))])
        # Crear la consulta SQL con los nombres de las columnas
        query = f"INSERT INTO {table_name} ({', '.join(column_names)}) VALUES ({placeholders})"
        # Ejecutar la consulta con los datos de la fila
        cursor.execute(query, row)

    # Confirmar cambios y cerrar la conexión
    conn.commit()
    conn.close()

if __name__ == "__main__":
    # Verificar que se proporcionen los argumentos adecuados
    if len(sys.argv) != 3:
        print("Usage: python script.py <file_path> <table_name>")
        sys.exit(1)

    # Obtener los argumentos del comando
    file_path = sys.argv[1]
    table_name = sys.argv[2]

    # Llamar a la función para insertar datos desde el archivo Excel
    insert_data_from_excel(file_path, table_name)
